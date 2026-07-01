import os
import re
import json
import random
import shutil
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

try:
    import requests
except ImportError:
    requests = None

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None


BASE_DIR = Path(__file__).resolve().parent.parent

if load_dotenv:
    load_dotenv(BASE_DIR / ".env")


def env_bool(nome, padrao=False):
    valor = os.getenv(nome)
    if valor is None:
        return padrao
    return str(valor).strip().lower() in {"1", "true", "sim", "s", "yes", "y", "on"}


def env_int(nome, padrao):
    try:
        return int(os.getenv(nome, str(padrao)))
    except ValueError:
        return padrao


BASE_PROCESSAMENTO = Path(
    os.getenv("BASE_PROCESSAMENTO", BASE_DIR / "examples")).resolve()
PASTA_IMPORTACAO_GERAL = Path(os.getenv(
    "PASTA_IMPORTACAO_GERAL", BASE_PROCESSAMENTO / "importacao_geral")).resolve()
MODELO_PLANILHA = Path(
    os.getenv("MODELO_PLANILHA", BASE_DIR / "modelo_importacao.xlsx")).resolve()

ARQUIVO_LOG = BASE_DIR / "log.txt"
ARQUIVO_CONTROLE = BASE_DIR / "copiados.txt"
ARQUIVO_DISPAROS = BASE_DIR / "disparos_usados.txt"
ARQUIVO_IMPORTADOS_API = BASE_DIR / "importados_api.txt"
PASTA_AUDITORIA_API = BASE_DIR / "auditoria_api"
PASTA_PAYLOADS_API = PASTA_AUDITORIA_API / "payloads"
PASTA_RESPOSTAS_API = PASTA_AUDITORIA_API / "respostas"
PASTA_ERRO_API = BASE_DIR / "erro_api"

TENTATIVAS_ABRIR_TXT = env_int("TENTATIVAS_ABRIR_TXT", 5)
ESPERA_ABRIR_TXT_SEGUNDOS = env_int("ESPERA_ABRIR_TXT_SEGUNDOS", 5)
HORARIO_ENCERRAMENTO = env_int("HORARIO_ENCERRAMENTO", 19)


# MAPAS GENÉRICOS DE EXEMPLO


MAPA_CARTEIRAS = {
    "clientea": "CLIENTE A",
    "clienteb": "CLIENTE B",
    "clientec": "CLIENTE C",
    "campanhaa": "CAMPANHA A",
    "campanhab": "CAMPANHA B",
}

MAPA_CONTRATANTES = {
    "clientea": "CONTRATANTE A",
    "clienteb": "CONTRATANTE B",
    "clientec": "CONTRATANTE C",
    "campanhaa": "CONTRATANTE A",
    "campanhab": "CONTRATANTE B",
}

MAPA_ETIQUETAS = {
    "clientea": "GRUPO A",
    "clienteb": "GRUPO B",
    "clientec": "GRUPO C",
    "campanhaa": "GRUPO A",
    "campanhab": "GRUPO B",
}

CAMPOS_OBRIGATORIOS = [
    "Nome", "Telefone", "Email", "Etiquetas", "Disparo", "Contratante",
    "Valor", "Data de vencimento", "CPF/CNPJ", "ID",
]


# CONTROLE E LOG

def log(msg):
    with open(ARQUIVO_LOG, "a", encoding="utf-8", errors="ignore") as f:
        data = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        f.write(f"{data} - {msg}\n")


def carregar_linhas(caminho):
    if not caminho.exists():
        return set()
    with open(caminho, "r", encoding="utf-8", errors="ignore") as f:
        return {l.strip() for l in f if l.strip()}


def adicionar_linha(caminho, valor):
    with open(caminho, "a", encoding="utf-8", errors="ignore") as f:
        f.write(str(valor) + "\n")


copiados = carregar_linhas(ARQUIVO_CONTROLE)


# CAMINHOS AUTOMÁTICOS POR DATA

def get_caminhos():
    hoje = datetime.now()
    ano = hoje.strftime("%Y")
    dia = hoje.strftime("%d.%m")
    mes = hoje.strftime("%m.%B")

    pasta_whats = BASE_PROCESSAMENTO / "WHATSAPP" / ano / mes / dia
    pasta_sms = BASE_PROCESSAMENTO / "SMS" / ano / mes / dia
    entrada_whats = pasta_whats / "entrada"
    entrada_sms = pasta_sms / "entrada"

    for pasta in [pasta_whats, pasta_sms, entrada_whats, entrada_sms, PASTA_IMPORTACAO_GERAL]:
        pasta.mkdir(parents=True, exist_ok=True)

    print(f"WHATS entrada: {entrada_whats}")
    print(f"SMS entrada: {entrada_sms}")
    return entrada_whats, entrada_sms, pasta_whats, pasta_sms


def abrir_txt_com_retentativas(caminho, nome_arquivo):
    for tentativa in range(1, TENTATIVAS_ABRIR_TXT + 1):
        try:
            return open(caminho, "r", encoding="latin-1", errors="ignore")
        except PermissionError as erro:
            log(f"ARQUIVO BLOQUEADO: {nome_arquivo} | tentativa {tentativa}/{TENTATIVAS_ABRIR_TXT} | {erro}")
            if tentativa < TENTATIVAS_ABRIR_TXT:
                time.sleep(ESPERA_ABRIR_TXT_SEGUNDOS)
    log(f"ARQUIVO IGNORADO TEMPORARIAMENTE: {nome_arquivo}")
    return None


# TRATAMENTOS

def limpar_texto(valor):
    return str(valor).replace('"', '').replace("'", "").strip()


def normalizar_telefone(tel):
    tel = ''.join(filter(str.isdigit, str(tel)))
    if tel.startswith("55") and len(tel) in [12, 13]:
        tel = tel[2:]
    if len(tel) not in [10, 11]:
        return None
    ddd = tel[:2]
    if not ddd.isdigit() or not (11 <= int(ddd) <= 99):
        return None
    return tel


def normalizar_telefone_sms(tel):
    tel = ''.join(filter(str.isdigit, str(tel)))
    if len(tel) in [10, 11] and not tel.startswith("55"):
        tel = "55" + tel
    if not tel.startswith("55") or len(tel) not in [12, 13]:
        return None
    ddd = tel[2:4]
    if not ddd.isdigit() or not (11 <= int(ddd) <= 99):
        return None
    return tel


def primeiro_nome(nome):
    nome = limpar_texto(nome)
    return nome.split()[0] if nome else ""


def remover_acentos(valor):
    texto = unicodedata.normalize("NFKD", str(valor))
    return "".join(c for c in texto if not unicodedata.combining(c))


def normalizar_busca(valor):
    texto = remover_acentos(valor).lower()
    return re.sub(r"[^a-z0-9]+", "", texto)


def carregar_disparos_usados():
    return carregar_linhas(ARQUIVO_DISPAROS)


def gerar_disparo_unico(carteira):
    usados = carregar_disparos_usados()
    data_atual = datetime.now().strftime("%d.%m.%Y")
    carteira_disparo = "_".join(str(carteira).strip().upper().split())
    while True:
        numero = random.randint(1000, 9999)
        disparo = f"{carteira_disparo}_{data_atual}_{numero}"
        if disparo not in usados:
            adicionar_linha(ARQUIVO_DISPAROS, disparo)
            return disparo


def identificar_info_arquivo(nome_arquivo):
    nome_busca = normalizar_busca(nome_arquivo)
    chave_encontrada = "padrao"
    for chave in MAPA_CARTEIRAS:
        if normalizar_busca(chave) in nome_busca:
            chave_encontrada = chave
            break
    carteira = MAPA_CARTEIRAS.get(chave_encontrada, "PADRAO")
    contratante = MAPA_CONTRATANTES.get(chave_encontrada, carteira)
    etiqueta = MAPA_ETIQUETAS.get(chave_encontrada, carteira.upper())
    disparo = gerar_disparo_unico(carteira)
    return disparo, contratante, carteira, etiqueta


# API EXTERNA OPCIONAL

def carregar_config_api():
    return {
        "enabled": env_bool("API_ENABLED", False),
        "token": os.getenv("API_TOKEN", "").strip(),
        "base_url": os.getenv("API_BASE_URL", "https://api.exemplo.com").strip().rstrip("/"),
        "timeout": env_int("API_TIMEOUT", 60),
        "dry_run": env_bool("API_DRY_RUN", True),
    }


def dividir_em_lotes(lista, tamanho=100):
    return [lista[i:i + tamanho] for i in range(0, len(lista), max(tamanho, 1))]


def montar_item_api(registro):
    return {
        "name": registro.get("Nome") or "",
        "phoneNumber": registro.get("Telefone") or "",
        "email": None,
        "tagNames": [registro.get("Etiquetas") or ""],
        "customFields": {
            "disparo": registro.get("Disparo") or "",
            "contratante": registro.get("Contratante") or "",
            "valor": registro.get("Valor") or "",
            "data_de_vencimento": registro.get("Data de vencimento") or "",
            "id": str(registro.get("ID") or ""),
        },
    }


def montar_payload_api(lote):
    return {
        "items": [montar_item_api(registro) for registro in lote],
        "options": {"upsert": True, "getIfExists": False},
    }


def nome_seguro_arquivo(valor):
    nome = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(valor))
    return nome.strip("._") or "arquivo"


def salvar_json(caminho, dados):
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with open(caminho, "w", encoding="utf-8", errors="ignore") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def enviar_lote_api(payload, numero_lote, config, nome_arquivo):
    quantidade = len(payload.get("items", []))
    nome_base = nome_seguro_arquivo(Path(nome_arquivo).stem)
    caminho_payload = PASTA_PAYLOADS_API / \
        f"{nome_base}_lote_{numero_lote}_{quantidade}_payload.json"
    salvar_json(caminho_payload, payload)

    if config["dry_run"]:
        log(f"API DRY RUN lote {numero_lote}: payload salvo em {caminho_payload}")
        print(
            f"[API DRY RUN] lote {numero_lote}: payload salvo em {caminho_payload}")
        return True

    if requests is None:
        log("API ERRO: biblioteca requests não instalada")
        return False
    if not config["token"]:
        log("API ERRO: API_TOKEN não configurado")
        return False

    url = config["base_url"] + "/contacts/batch"
    headers = {"Authorization": "Bearer " +
               config["token"], "Content-Type": "application/json"}

    try:
        resposta = requests.post(
            url, json=payload, headers=headers, timeout=config["timeout"])
        caminho_resposta = PASTA_RESPOSTAS_API / \
            f"{nome_base}_lote_{numero_lote}_{quantidade}_response.txt"
        caminho_resposta.parent.mkdir(parents=True, exist_ok=True)
        caminho_resposta.write_text(
            resposta.text, encoding="utf-8", errors="ignore")
        log(f"API lote {numero_lote}: status={resposta.status_code} | resposta={caminho_resposta}")
        return 200 <= resposta.status_code < 300
    except requests.RequestException as erro:
        log(f"API ERRO lote {numero_lote}: {erro}")
        return False


def enviar_contatos_api(dados, nome_arquivo):
    config = carregar_config_api()
    if not config["enabled"]:
        return True
    for indice, lote in enumerate(dividir_em_lotes(dados, 100), start=1):
        if not enviar_lote_api(montar_payload_api(lote), indice, config, nome_arquivo):
            return False
        time.sleep(1)
    adicionar_linha(ARQUIVO_IMPORTADOS_API,
                    f"{nome_arquivo};{datetime.now():%Y-%m-%d %H:%M:%S}")
    return True


# EXCEL / WHATS

def localizar_colunas(ws):
    colunas = {str(cell.value).strip()               : cell.column for cell in ws[1] if cell.value}
    faltando = [campo for campo in CAMPOS_OBRIGATORIOS if campo not in colunas]
    if faltando:
        raise ValueError(
            "A planilha modelo está sem estes cabeçalhos: " + ", ".join(faltando))
    return colunas


def processar_arquivo_whats(caminho, nome_arquivo, pasta_saida):
    dados, vistos, ignorados = [], set(), 0
    arquivo_txt = abrir_txt_com_retentativas(caminho, nome_arquivo)
    if arquivo_txt is None:
        return

    disparo, contratante, carteira, etiqueta = identificar_info_arquivo(
        nome_arquivo)

    with arquivo_txt as f:
        for num_linha, linha in enumerate(f, start=1):
            partes = [limpar_texto(p) for p in limpar_texto(linha).split(";")]
            if len(partes) < 3:
                ignorados += 1
                log(f"{nome_arquivo} linha {num_linha} ignorada: menos de 3 campos")
                continue

            telefone = normalizar_telefone(partes[0])
            if not telefone:
                ignorados += 1
                log(f"{nome_arquivo} linha {num_linha} ignorada: telefone inválido")
                continue
            if telefone in vistos:
                continue
            vistos.add(telefone)

            nome = primeiro_nome(partes[1])
            if not nome:
                ignorados += 1
                log(f"{nome_arquivo} linha {num_linha} ignorada: nome vazio")
                continue

            dados.append({
                "Nome": nome,
                "Telefone": telefone,
                "Email": None,
                "Etiquetas": etiqueta,
                "Disparo": disparo,
                "Contratante": contratante,
                "Valor": partes[3] if len(partes) >= 4 else None,
                "Data de vencimento": partes[4] if len(partes) >= 5 else None,
                "CPF/CNPJ": None,
                "ID": partes[2],
            })

    if not dados:
        print(f"[WHATS] {nome_arquivo}: nenhum registro válido encontrado")
        return

    temp = BASE_DIR / "temp_whats.xlsx"
    if temp.exists():
        temp.unlink()
    shutil.copy(MODELO_PLANILHA, temp)

    wb = load_workbook(temp)
    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    colunas = localizar_colunas(ws)

    for linha_excel, registro in enumerate(dados, start=2):
        for campo, valor in registro.items():
            ws.cell(linha_excel, colunas[campo], valor)
    for linha_excel in range(2, ws.max_row + 1):
        ws.cell(linha_excel, colunas["Telefone"]).number_format = "@"
        ws.cell(linha_excel, colunas["ID"]).number_format = "@"

    nome_saida = Path(nome_arquivo).with_suffix(".xlsx").name
    destino_xlsx = pasta_saida / nome_saida
    wb.save(temp)
    shutil.move(temp, destino_xlsx)

    if not enviar_contatos_api(dados, nome_arquivo):
        PASTA_ERRO_API.mkdir(parents=True, exist_ok=True)
        shutil.move(caminho, PASTA_ERRO_API / nome_arquivo)
        print(f"[WHATS] {nome_arquivo}: XLSX gerado, mas API falhou")
        return

    shutil.move(caminho, pasta_saida / nome_arquivo)
    print(
        f"[WHATS] {nome_arquivo} -> {nome_saida} | {len(dados)} registros | {ignorados} ignorados")
    log(f"{nome_arquivo} WHATS OK ({len(dados)} registros, {ignorados} ignorados)")


# SMS / TXT

def processar_sms_txt(caminho, nome_arquivo, pasta_saida):
    linhas = ["TELEFONE;NOME;IDCLIENTE"]
    ignorados = 0
    with open(caminho, "r", encoding="latin-1", errors="ignore") as f:
        for num_linha, linha in enumerate(f, start=1):
            partes = [limpar_texto(p) for p in limpar_texto(linha).split(";")]
            if num_linha == 1 and len(partes) >= 3 and partes[0].upper() == "TELEFONE":
                continue
            if len(partes) < 3:
                ignorados += 1
                continue
            telefone = normalizar_telefone_sms(partes[0])
            nome = partes[1].strip().upper()
            id_cliente = partes[2].strip()
            if not telefone or not nome or not id_cliente:
                ignorados += 1
                continue
            linhas.append(f"{telefone};{nome};{id_cliente}")

    destino = pasta_saida / nome_arquivo
    destino.write_text("\n".join(linhas) + "\n",
                       encoding="utf-8", errors="ignore")
    os.remove(caminho)
    print(
        f"[SMS] {nome_arquivo} -> TXT editado | {len(linhas) - 1} registros | {ignorados} ignorados")
    log(f"{nome_arquivo} SMS TXT OK ({len(linhas) - 1} registros, {ignorados} ignorados)")


# CÓPIA PARA IMPORTAÇÃO GERAL

def copiar_feitos(pasta_whats, pasta_sms):
    global copiados
    for pasta in [pasta_whats / "feitos", pasta_sms / "feitos"]:
        if not pasta.exists():
            continue
        for origem in pasta.glob("*.txt"):
            if origem.name in copiados or origem.stat().st_size == 0:
                continue
            destino = PASTA_IMPORTACAO_GERAL / origem.name
            shutil.copy(origem, destino)
            adicionar_linha(ARQUIVO_CONTROLE, origem.name)
            copiados.add(origem.name)
            print(f"Copiado uma vez: {origem.name}")
            log(f"Copiado para importação geral: {origem.name}")


# INÍCIO

def main():
    entrada_whats, entrada_sms, pasta_whats, pasta_sms = get_caminhos()
    print(f"Automação rodando até {HORARIO_ENCERRAMENTO}h...")

    while datetime.now().hour < HORARIO_ENCERRAMENTO:
        for arquivo in entrada_whats.glob("*.txt"):
            processar_arquivo_whats(arquivo, arquivo.name, pasta_whats)
        for arquivo in entrada_sms.glob("*.txt"):
            processar_sms_txt(arquivo, arquivo.name, pasta_sms)
        copiar_feitos(pasta_whats, pasta_sms)
        time.sleep(30)

    print("Automação encerrada")


if __name__ == "__main__":
    main()
